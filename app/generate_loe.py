#!/usr/bin/env python3
""" LoE generator application """

# Imports - Python Standard Library
from pathlib import Path
from os import listdir, unlink
from os.path import join
from time import ctime
from typing import Dict, List

# Imports - Third-Party
import openpyxl
from yaml import safe_load

# Imports - Local

# Constants
CURRENT_DIR = Path(__file__).parent
DATA_DIR = 'data'
SPREADSHEET_NAME = f'LoE - {ctime()}.xlsx'
SPREADSHEET_PATH = join(CURRENT_DIR, DATA_DIR, SPREADSHEET_NAME)
SOURCE_DATA_DIR = 'source'
SOURCE_DATA_FILE_NAME = 'source_data.yml'
SOURCE_DATA_FILE_PATH = join(CURRENT_DIR, SOURCE_DATA_DIR)


def clear_workbook_dir() -> List:
    """ Remove existing Excel workbooks.

        Args:
            None.

        Returns:
            data_dir_files (List):
                List of files in the data file directory.
    """

    # Set the path to the data file directory
    data_dir = Path(SPREADSHEET_PATH).parent

    # Create a list of the files in the data file directory
    data_dir_files = listdir(data_dir)

    # Remove/unlink the files in the data file directory
    for file in data_dir_files:
        unlink(
            path=join(data_dir, file)
        )

    # Collect a new list of the files in the data file directory
    data_dir_files = listdir(data_dir)

    return data_dir_files


def create_new_workbook() -> None:
    """ Create a new Excel workbook.

        Args:
            None.

        Returns:
            wb_name (str):
                Path to the new spreadsheet file.
    """

    # Create a Workbook object
    wb = openpyxl.Workbook()

    # Add a title to the first/default worksheet
    ws = wb.active
    ws.title = 'LoE #1'

    # Save the Workbook to a file
    wb_name = join(SPREADSHEET_PATH)
    wb.save(
        filename=wb_name
    )

    return wb_name


def read_source_data(
    data_file: str = SOURCE_DATA_FILE_NAME
) -> Dict:
    """ Read source YAML file.

        Args:
            data_file (str):
                Name of the source data YAML file.

        Returns:
            project_data (Dict):
                Project source data as dictionary data.
    """

    with open(
        file=join(SOURCE_DATA_FILE_PATH, data_file),
        mode='rt',
        encoding='utf-8'
    ) as source_data:

        project_data = safe_load(
            stream=source_data
        )

    return project_data


def create_loe(
    project_data: Dict
) -> None:
    """ Create an LoE Excel spreadsheet with source data.

        Args:
            project_data (Dict):
                Project data imported from a YAML file.

        Returns:
            None.
    """

    # Create a new Excel spreadsheet
    wb_name = create_new_workbook()

    # Load the workbook file
    wb = openpyxl.load_workbook(
        filename=wb_name
    )

    # Assign the first worksheet to a variable
    ws_1 = wb[wb.worksheets[0]]

    # Set the project name in the first cell
    ws_1.cell(
        row=1,
        column=1
    ).value = project_data['project']['name']

    # Set the project client in the second cell
    ws_1.cell(
        row=2,
        column=1
    ).value = project_data['project']['name']

    # Set the project description in the third cell
    ws_1.cell(
        row=3,
        column=1
    ).value = project_data['project']['description']

    return None


def main() -> None:
    """ Main application. """

    return None


if __name__ == '__main__':
    main()
